import os
import uuid
import pandas as pd
import threading
import time
import ipaddress
from flask import Flask, request, render_template, send_file, redirect, url_for, flash
from werkzeug.utils import secure_filename
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = 'secure_key_for_flash_messages'

UPLOAD_FOLDER = 'temp_uploads'
OUTPUT_FOLDER = 'temp_outputs'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

@app.route('/api/system/restart', methods=['POST'])
def system_restart():
    if request.remote_addr != '127.0.0.1':
        return "Unauthorized", 403
    def kill_app():
        time.sleep(1)
        os._exit(0)
    threading.Thread(target=kill_app).start()
    return "Application is restarting to apply new code...", 200

# -------------------------------------------------------------------
# SHADOWED RULES DETECTION ENGINE (Phase 3 Analytics)
# -------------------------------------------------------------------
def detect_shadowed_rules(df, address_column):
    """
    Detects shadowed rules using subnet math (ipaddress library).
    A rule is shadowed if its IP/subnet is covered by a higher-priority rule.
    Gracefully handles non-IP values (e.g., object names) by catching ValueError.
    Returns a DataFrame of shadowed rules.
    """
    if address_column not in df.columns:
        return pd.DataFrame(columns=df.columns)
    
    shadowed_indices = []
    
    for i in range(len(df)):
        current_addr = str(df.iloc[i][address_column]).strip()
        
        # Skip empty or 'any' addresses
        if not current_addr or current_addr.lower() in ['any', '']:
            continue
        
        try:
            # Try to parse current rule's address as network/IP
            current_net = ipaddress.ip_network(current_addr, strict=False)
        except (ValueError, ipaddress.AddressValueError, ipaddress.NetmaskValueError):
            # Not an IP/subnet, skip
            continue
        
        # Check all higher-priority rules (rules above current rule)
        is_shadowed = False
        for j in range(i):
            higher_addr = str(df.iloc[j][address_column]).strip()
            
            if not higher_addr or higher_addr.lower() in ['any', '']:
                continue
            
            try:
                higher_net = ipaddress.ip_network(higher_addr, strict=False)
                # If current rule's subnet is within higher rule's subnet, it's shadowed
                if current_net.subnet_of(higher_net):
                    is_shadowed = True
                    break
            except (ValueError, ipaddress.AddressValueError, ipaddress.NetmaskValueError):
                # Not an IP/subnet, skip
                continue
        
        if is_shadowed:
            shadowed_indices.append(i)
    
    return df.iloc[shadowed_indices].copy() if shadowed_indices else pd.DataFrame(columns=df.columns)

# -------------------------------------------------------------------
# ENGINE 1: PALO ALTO PROCESSOR
# -------------------------------------------------------------------
def process_palo_alto_rules(df, output_path, site_name, nist_metrics=None, cis_metrics=None):
    disabled_df = df[df['Name'].astype(str).str.contains('Disabled', case=False, na=False)]
    active_df = df[~df['Name'].astype(str).str.contains('Disabled', case=False, na=False)]
    
    hit_col = 'Rule Usage Hit Count' if 'Rule Usage Hit Count' in df.columns else 'Rule Usage Rule Usage'
    zero_hit_df = active_df[
        (active_df[hit_col].astype(str).str.strip().str.lower() == 'unused') | 
        (active_df[hit_col] == 0) | (active_df[hit_col] == '0')
    ]
    
    active_hit_df = active_df.drop(zero_hit_df.index)
    
    # ---------------------------------------------------------
    # PHASE 1: SPLIT PROTOCOL FLAGGING (CSV Compatible)
    # ---------------------------------------------------------
    admin_pattern = r'(?i)(?:^|[-_\s])(?:22|3389|rdp|ssh)(?:[-_\s]|$)'
    cleartext_pattern = r'(?i)(?:^|[-_\s])(?:21|23|80|ftp|telnet|http)(?:[-_\s]|$)'
    
    admin_ports_df = active_hit_df[
        active_hit_df['Service'].astype(str).str.contains(admin_pattern, regex=True, na=False) & 
        (active_hit_df['Action'].astype(str).str.strip().str.lower().isin(['allow', 'accept']))
    ]
    
    cleartext_ports_df = active_hit_df[
        active_hit_df['Service'].astype(str).str.contains(cleartext_pattern, regex=True, na=False) & 
        (active_hit_df['Action'].astype(str).str.strip().str.lower().isin(['allow', 'accept']))
    ]

    # ---------------------------------------------------------
    # PHASE 2: BI-DIRECTIONAL & ZERO HIT CHECK (CSV Compatible)
    # ---------------------------------------------------------
    # 1. Search across ALL active rules, but exclude 'any' IPs and 'deny/drop' rules
    valid_bidir = active_df[
        (~active_df['Source Address'].astype(str).str.strip().str.lower().isin(['any', ''])) &
        (~active_df['Destination Address'].astype(str).str.strip().str.lower().isin(['any', ''])) &
        (active_df['Action'].astype(str).str.strip().str.lower().isin(['allow', 'accept'])) # NEW: Only check allow rules!
    ]
    
    # 2. Self-Merge the dataframe where Rule A Src == Rule B Dst AND Rule A Dst == Rule B Src
    merged_bidir = pd.merge(
        valid_bidir, valid_bidir, 
        left_on=['Source Address', 'Destination Address', 'Service'], 
        right_on=['Destination Address', 'Source Address', 'Service'], 
        suffixes=('_A', '_B')
    )
    
    # 3. Filter out rows matching against themselves, and extract unique rule names
    merged_bidir = merged_bidir[merged_bidir['Name_A'] != merged_bidir['Name_B']]
    bidirectional_names = merged_bidir['Name_A'].unique()
    
    # 4. Only flag the mirrored rule if it ALSO has Zero Hits!
    bidirectional_df = zero_hit_df[zero_hit_df['Name'].isin(bidirectional_names)]
    
    # ---------------------------------------------------------
    # PHASE 3: SHADOWED RULES DETECTION (Phase 3 Analytics)
    # ---------------------------------------------------------
    shadowed_src_df = detect_shadowed_rules(active_df, 'Source Address')
    shadowed_dst_df = detect_shadowed_rules(active_df, 'Destination Address')
    shadowed_df = pd.concat([shadowed_src_df, shadowed_dst_df]).drop_duplicates(subset=['Name'], keep='first')
    
    # ---------------------------------------------------------
    # EXISTING COMPLIANCE CHECKS
    # ---------------------------------------------------------
    if 'NIST Documented' in active_hit_df.columns:
        undocumented_df = active_hit_df[active_hit_df['NIST Documented'].astype(str).str.strip().str.lower() == 'no']
    else:
        undocumented_df = pd.DataFrame()
    
    source_any_df = active_hit_df[active_hit_df['Source Address'].astype(str).str.strip().str.lower() == 'any']
    dest_any_df = active_hit_df[
        (active_hit_df['Destination Address'].astype(str).str.strip().str.lower() == 'any') & 
        (active_hit_df['Application'].astype(str).str.strip().str.lower() == 'any') & 
        (active_hit_df['URL Category'].astype(str).str.strip().str.lower() == 'any')
    ]
    service_any_df = active_hit_df[
        (active_hit_df['Service'].astype(str).str.strip().str.lower().isin(['any', 'application-default'])) & 
        (active_hit_df['Application'].astype(str).str.strip().str.lower() == 'any') & 
        (active_hit_df['URL Category'].astype(str).str.strip().str.lower() == 'any')
    ]
    
    logs_none_df = active_hit_df[~active_hit_df['Options'].astype(str).str.contains('Log Forwarding Profile setting', case=False, na=False)]
    tags_none_df = active_hit_df[active_hit_df['Tags'].astype(str).str.strip().str.lower() == 'none']
    profile_none_df = active_hit_df[active_hit_df['Profile'].astype(str).str.strip().str.lower() == 'none']

    is_src_any = active_df['Source Address'].astype(str).str.strip().str.lower() == 'any'
    is_dst_any = (active_df['Destination Address'].astype(str).str.strip().str.lower() == 'any') & (active_df['Application'].astype(str).str.strip().str.lower() == 'any') & (active_df['URL Category'].astype(str).str.strip().str.lower() == 'any')
    is_srv_any = (active_df['Service'].astype(str).str.strip().str.lower().isin(['any', 'application-default'])) & (active_df['Application'].astype(str).str.strip().str.lower() == 'any') & (active_df['URL Category'].astype(str).str.strip().str.lower() == 'any')
    is_prof_none = active_df['Profile'].astype(str).str.strip().str.lower() == 'none'
    
    # Flag Cleartext as High Risk, Admin Ports as Medium Risk
    is_cleartext = active_df['Service'].astype(str).str.contains(cleartext_pattern, regex=True, na=False) & (active_df['Action'].astype(str).str.strip().str.lower().isin(['allow', 'accept']))
    is_admin = active_df['Service'].astype(str).str.contains(admin_pattern, regex=True, na=False) & (active_df['Action'].astype(str).str.strip().str.lower().isin(['allow', 'accept']))
    
    is_zero_hit = (active_df[hit_col].astype(str).str.strip().str.lower() == 'unused') | (active_df[hit_col] == 0) | (active_df[hit_col] == '0')
    is_log_none = ~active_df['Options'].astype(str).str.contains('Log Forwarding Profile setting', case=False, na=False)
    is_tag_none = active_df['Tags'].astype(str).str.strip().str.lower() == 'none'
    
    high_mask = is_src_any | is_dst_any | is_srv_any | is_prof_none | is_cleartext
    med_mask = (~high_mask) & (is_zero_hit | is_log_none | is_tag_none | is_admin)
    low_mask = (~high_mask) & (~med_mask)

    metrics = {
        'total_rules': len(df), 'disabled_rules': len(disabled_df), 'active_rules': len(active_df),
        'zero_hit_rules': len(zero_hit_df), 'active_hit_rules': len(active_hit_df),
        'source_any': len(source_any_df), 'destination_any': len(dest_any_df),
        'service_any': len(service_any_df), 'profile_issue': len(profile_none_df),
        'logs_none': len(logs_none_df), 'tags_none': len(tags_none_df),
        'admin_ports': len(admin_ports_df),
        'cleartext_ports': len(cleartext_ports_df),
        'bidirectional_rules': len(bidirectional_df),
        'undocumented_rules': len(undocumented_df),
        'shadowed_rules': len(shadowed_df),
        'high_risk': int(high_mask.sum()), 'medium_risk': int(med_mask.sum()), 'low_risk': int(low_mask.sum())
    }

    generate_excel_report(output_path, site_name, metrics, df, disabled_df, active_df, zero_hit_df, active_hit_df, source_any_df, dest_any_df, service_any_df, profile_none_df, logs_none_df, tags_none_df, admin_ports_df, cleartext_ports_df, bidirectional_df, undocumented_df, shadowed_df, nist_metrics, cis_metrics)

    web_cols = ['Name', 'Source Zone', 'Source Address', 'Destination Zone', 'Destination Address', 'Application', 'Service', 'Action', 'NIST Documented']
    web_cols = [col for col in web_cols if col in df.columns] 
    
    web_data = {
        'total': df[web_cols].to_dict('records'),
        'disabled': disabled_df[web_cols].to_dict('records'),
        'active': active_df[web_cols].to_dict('records'),
        'high_risk': active_df[high_mask][web_cols].to_dict('records'),
        'med_risk': active_df[med_mask][web_cols].to_dict('records'),
        'low_risk': active_df[low_mask][web_cols].to_dict('records'),
        'zero_hit': zero_hit_df[web_cols].to_dict('records'),
        'active_hit': active_hit_df[web_cols].to_dict('records'),
        'source_any': source_any_df[web_cols].to_dict('records'),
        'dest_any': dest_any_df[web_cols].to_dict('records'),
        'service_any': service_any_df[web_cols].to_dict('records'),
        'profile_issue': profile_none_df[web_cols].to_dict('records'),
        'logs_none': logs_none_df[web_cols].to_dict('records'),
        'tags_none': tags_none_df[web_cols].to_dict('records'),
        'admin_ports': admin_ports_df[web_cols].to_dict('records'),
        'cleartext_ports': cleartext_ports_df[web_cols].to_dict('records'),
        'bidirectional': bidirectional_df[web_cols].to_dict('records'),
        'undocumented': undocumented_df[web_cols].to_dict('records') if not undocumented_df.empty else [],
        'shadowed': shadowed_df[web_cols].to_dict('records') if not shadowed_df.empty else [],
        'nist': nist_metrics or {},
        'cis': cis_metrics or {}
    }
    return metrics, web_data
# -------------------------------------------------------------------
# ENGINE 2: FORTINET PROCESSOR
# -------------------------------------------------------------------
def process_fortinet_rules(df, output_path, site_name):
    disabled_df = df[df['Status'].astype(str).str.strip().str.lower() == 'disabled']
    active_df = df[df['Status'].astype(str).str.strip().str.lower() == 'enabled']
    
    zero_hit_df = active_df[(active_df['Hit Count'] == 0) | (active_df['Hit Count'].astype(str).str.strip() == '0')]
    active_hit_df = active_df.drop(zero_hit_df.index)
    
    source_any_df = active_hit_df[active_hit_df['Source'].astype(str).str.strip().str.lower() == 'all']
    dest_any_df = active_hit_df[active_hit_df['Destination'].astype(str).str.strip().str.lower() == 'all']
    service_any_df = active_hit_df[active_hit_df['Service'].astype(str).str.strip().str.lower() == 'all']
    
    profile_none_df = active_hit_df[
        (active_hit_df['Action'].astype(str).str.strip().str.upper() == 'ACCEPT') & 
        ((active_hit_df['Security Profiles'] == '') | (active_hit_df['Security Profiles'].astype(str).str.contains('no-inspection', case=False)))
    ]
    
    logs_none_df = active_hit_df[
        (active_hit_df['Log'] == '') | 
        (active_hit_df['Log'].astype(str).str.strip().str.lower() == 'disabled') |
        (active_hit_df['Log'].astype(str).str.strip().str.lower() == 'utm')
    ]
    
    tags_none_df = pd.DataFrame(columns=df.columns)

    # ---------------------------------------------------------
    # PHASE 3: SHADOWED RULES DETECTION (Phase 3 Analytics)
    # ---------------------------------------------------------
    shadowed_src_df = detect_shadowed_rules(active_df, 'Source')
    shadowed_dst_df = detect_shadowed_rules(active_df, 'Destination')
    shadowed_combined = pd.concat([shadowed_src_df, shadowed_dst_df])
    # Get unique rule identifier - use 'Policy' if it exists, otherwise 'Name'
    rule_id_col = 'Policy' if 'Policy' in active_df.columns else 'Name'
    shadowed_df = shadowed_combined.drop_duplicates(subset=[rule_id_col], keep='first') if not shadowed_combined.empty else pd.DataFrame(columns=df.columns)

    is_accept = active_df['Action'].astype(str).str.strip().str.upper() == 'ACCEPT'
    is_src_all = active_df['Source'].astype(str).str.strip().str.lower() == 'all'
    is_dst_all = active_df['Destination'].astype(str).str.strip().str.lower() == 'all'
    is_srv_all = active_df['Service'].astype(str).str.strip().str.lower() == 'all'
    is_prof_none = is_accept & ((active_df['Security Profiles'] == '') | (active_df['Security Profiles'].astype(str).str.contains('no-inspection', case=False)))
    is_zero_hit = (active_df['Hit Count'] == 0) | (active_df['Hit Count'].astype(str).str.strip() == '0')
    
    is_log_disabled = (
        (active_df['Log'] == '') | 
        (active_df['Log'].astype(str).str.strip().str.lower() == 'disabled') |
        (active_df['Log'].astype(str).str.strip().str.lower() == 'utm')
    )
    
    high_mask = is_accept & (is_src_all | is_dst_all | is_srv_all | is_prof_none)
    med_mask = (~high_mask) & (is_zero_hit | is_log_disabled)
    low_mask = (~high_mask) & (~med_mask)

    metrics = {
        'total_rules': len(df), 'disabled_rules': len(disabled_df), 'active_rules': len(active_df),
        'zero_hit_rules': len(zero_hit_df), 'active_hit_rules': len(active_hit_df),
        'source_any': len(source_any_df), 'destination_any': len(dest_any_df),
        'service_any': len(service_any_df), 'profile_issue': len(profile_none_df),
        'logs_none': len(logs_none_df), 'tags_none': 0,
        'shadowed_rules': len(shadowed_df),
        'high_risk': int(high_mask.sum()), 'medium_risk': int(med_mask.sum()), 'low_risk': int(low_mask.sum())
    }

    generate_excel_report(output_path, site_name, metrics, df, disabled_df, active_df, zero_hit_df, active_hit_df, source_any_df, dest_any_df, service_any_df, profile_none_df, logs_none_df, tags_none_df, shadowed_df=shadowed_df)

    web_df = df.copy()
    web_df['Name'] = web_df['Name'] if 'Name' in web_df.columns else web_df['Policy']
    web_df['Source Zone'] = web_df['From']
    web_df['Source Address'] = web_df['Source']
    web_df['Destination Zone'] = web_df['To']
    web_df['Destination Address'] = web_df['Destination']
    web_df['Application'] = 'N/A' 
    web_df['Service'] = web_df['Service']
    web_df['Action'] = web_df['Action']
    
    web_cols = ['Name', 'Source Zone', 'Source Address', 'Destination Zone', 'Destination Address', 'Application', 'Service', 'Action']
    
    # NEW: Expanded web_data to include all dashboard metrics
    web_data = {
        'total': web_df[web_cols].to_dict('records'),
        'disabled': web_df.loc[disabled_df.index][web_cols].to_dict('records'),
        'active': web_df.loc[active_df.index][web_cols].to_dict('records'),
        'high_risk': web_df.loc[active_df[high_mask].index][web_cols].to_dict('records'),
        'med_risk': web_df.loc[active_df[med_mask].index][web_cols].to_dict('records'),
        'low_risk': web_df.loc[active_df[low_mask].index][web_cols].to_dict('records'),
        'zero_hit': web_df.loc[zero_hit_df.index][web_cols].to_dict('records'),
        'active_hit': web_df.loc[active_hit_df.index][web_cols].to_dict('records'),
        'source_any': web_df.loc[source_any_df.index][web_cols].to_dict('records'),
        'dest_any': web_df.loc[dest_any_df.index][web_cols].to_dict('records'),
        'service_any': web_df.loc[service_any_df.index][web_cols].to_dict('records'),
        'profile_issue': web_df.loc[profile_none_df.index][web_cols].to_dict('records'),
        'logs_none': web_df.loc[logs_none_df.index][web_cols].to_dict('records'),
        'tags_none': web_df.loc[tags_none_df.index][web_cols].to_dict('records'),
        'shadowed': web_df.loc[shadowed_df.index][web_cols].to_dict('records') if not shadowed_df.empty else []
    }
    return metrics, web_data


# -------------------------------------------------------------------
# SHARED EXCEL GENERATOR
# -------------------------------------------------------------------
def generate_excel_report(output_path, site_name, metrics, df, disabled_df, active_df, zero_hit_df, active_hit_df, source_any_df, dest_any_df, service_any_df, profile_none_df, logs_none_df, tags_none_df, admin_ports_df=None, cleartext_ports_df=None, bidirectional_df=None, undocumented_df=None, shadowed_df=None, nist_metrics=None, cis_metrics=None):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Build Standard Dashboard Data
        dashboard_data = [
            [site_name.upper(), "", ""],
            ["Total Rules", "Disabled Rules", "Active Rules"],
            [metrics.get('total_rules', 0), metrics.get('disabled_rules', 0), metrics.get('active_rules', 0)],
            ["High Risk", "Medium Risk", "Low Risk"],
            [metrics.get('high_risk', 0), metrics.get('medium_risk', 0), metrics.get('low_risk', 0)],
            ["Active Rules", "Zero Hit Review", "Active Hit Rules"],
            [metrics.get('active_rules', 0), metrics.get('zero_hit_rules', 0), metrics.get('active_hit_rules', 0)],
            ["Active Hit Rules", "Source Any", "To be reviewed"],
            [metrics.get('active_hit_rules', 0), metrics.get('source_any', 0), metrics.get('source_any', 0)],
            ["Active Hit Rules", "Destination Any", "To be reviewed"],
            [metrics.get('active_hit_rules', 0), metrics.get('destination_any', 0), metrics.get('destination_any', 0)],
            ["Active Hit Rules", "Service Any", "To be reviewed"],
            [metrics.get('active_hit_rules', 0), metrics.get('service_any', 0), metrics.get('service_any', 0)],
            
            # Phase 1 & 2 Dashboard Inserts
            ["Active Hit Rules", "Cleartext Protocols", "To be reviewed"],
            [metrics.get('active_hit_rules', 0), metrics.get('cleartext_ports', 0), metrics.get('cleartext_ports', 0)],
            ["Active Hit Rules", "Admin Mgmt Ports", "To be reviewed"],
            [metrics.get('active_hit_rules', 0), metrics.get('admin_ports', 0), metrics.get('admin_ports', 0)],
            ["Active Hit Rules", "Bi-Directional Duplicates", "To be reviewed"],
            [metrics.get('active_hit_rules', 0), metrics.get('bidirectional_rules', 0), metrics.get('bidirectional_rules', 0)],
            
            ["Active Rules", "Shadowed Rules (Phase 3)", "To be reviewed"],
            [metrics.get('active_rules', 0), metrics.get('shadowed_rules', 0), metrics.get('shadowed_rules', 0)],
            
            ["Active Hit Rules", "NIST Violations (Undoc)", "To be reviewed"],
            [metrics.get('active_hit_rules', 0), metrics.get('undocumented_rules', 0), metrics.get('undocumented_rules', 0)],
            ["Active Hit Rules", "Profile Issue", ""],
            [metrics.get('active_hit_rules', 0), metrics.get('profile_issue', 0), ""],
            ["Active Hit Rules", "Logs None", ""],
            [metrics.get('active_hit_rules', 0), metrics.get('logs_none', 0), ""],
            ["Active Hit Rules", "Tags None", ""],
            [metrics.get('active_hit_rules', 0), metrics.get('tags_none', 0), ""]
        ]

        if nist_metrics:
            dashboard_data.extend([
                ["", "", ""],
                ["NIST 800-41 TELEMETRY", "", ""],
                ["Firmware Version", nist_metrics.get('sw_version', 'N/A'), ""],
                ["Default Deny Enforced", "Yes" if nist_metrics.get('default_deny_enforced') else "No", ""]
            ])
            insecure = nist_metrics.get('insecure_mgt_profiles', [])
            if insecure:
                dashboard_data.append(["Insecure Mgmt Profiles", ", ".join(insecure), ""])
            else:
                dashboard_data.append(["Insecure Mgmt Profiles", "None (Secured)", ""])

        if cis_metrics:
            dashboard_data.extend([
                ["", "", ""],
                ["CIS PAN-OS BENCHMARKS", "", ""]
            ])
            for key, data in cis_metrics.items():
                dashboard_data.append([data['desc'], data['status'], data['value']])

        dashboard_df = pd.DataFrame(dashboard_data)
        dashboard_df.to_excel(writer, sheet_name='Dashboard', index=False, header=False)
        
        # Write Standard Sheets
        df.to_excel(writer, sheet_name='Master Sheet', index=False)
        disabled_df.to_excel(writer, sheet_name='Disabled Rules', index=False)
        active_df.to_excel(writer, sheet_name='Active Rules', index=False)
        zero_hit_df.to_excel(writer, sheet_name='Zero Hit Rules', index=False)
        active_hit_df.to_excel(writer, sheet_name='Active Hit Rules', index=False)
        
        # Write Phase 1 & 2 Sheets
        if admin_ports_df is not None: admin_ports_df.to_excel(writer, sheet_name='Admin Ports', index=False)
        if cleartext_ports_df is not None: cleartext_ports_df.to_excel(writer, sheet_name='Cleartext Protocols', index=False)
        if bidirectional_df is not None: bidirectional_df.to_excel(writer, sheet_name='Bi-Directional Duplicates', index=False)
        if shadowed_df is not None and not shadowed_df.empty: shadowed_df.to_excel(writer, sheet_name='Shadowed Rules', index=False)
        if undocumented_df is not None: undocumented_df.to_excel(writer, sheet_name='NIST Violations', index=False)
            
        source_any_df.to_excel(writer, sheet_name='Source Any Rules', index=False)
        dest_any_df.to_excel(writer, sheet_name='Destination Any Rules', index=False)
        service_any_df.to_excel(writer, sheet_name='Service Any Rules', index=False)
        profile_none_df.to_excel(writer, sheet_name='Profile None', index=False)
        logs_none_df.to_excel(writer, sheet_name='Logs None', index=False)
        tags_none_df.to_excel(writer, sheet_name='Tags None', index=False)
        
        # ... [Leave the rest of the dynamic Excel formatting block exactly as it is]
        # Apply Dynamic Excel Styling
        workbook = writer.book
        ws = writer.sheets['Dashboard']
        header_fill = PatternFill(start_color="2A3F54", end_color="2A3F54", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws['A1'].font = Font(size=14, bold=True) 
        
        # Dynamically color the rows so we don't have to hardcode header numbers!
        for row_idx, row_data in enumerate(dashboard_data, start=1):
            if row_data[0] == "": # Skip spacing rows
                continue
                
            for col_idx, col_letter in enumerate(['A', 'B', 'C']):
                cell = ws[f"{col_letter}{row_idx}"]
                cell.alignment = center_align
                if row_idx > 1: # Don't put a tight border on the Site Name
                    cell.border = thin_border
                    
            # Color header rows (even rows up to 22, plus the NIST title)
            is_standard_header = (row_idx % 2 == 0 and row_idx <= 22)
            is_nist_title = (row_data[0] == "NIST 800-41 TELEMETRY")
            
            if is_standard_header or is_nist_title:
                for col_letter in ['A', 'B', 'C']:
                    ws[f"{col_letter}{row_idx}"].fill = header_fill
                    ws[f"{col_letter}{row_idx}"].font = header_font

# -------------------------------------------------------------------
# ROUTING & AUTO-DETECTION
# -------------------------------------------------------------------
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        site_name = request.form.get('site_name', 'UNSPECIFIED TARGET').strip()
        if 'file' not in request.files:
            flash('No file uploaded.')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No selected file.')
            return redirect(request.url)

        if file:
            filename = secure_filename(file.filename)
            unique_id = str(uuid.uuid4())[:8]
            upload_path = os.path.join(UPLOAD_FOLDER, f"{unique_id}_{filename}")
            file.save(upload_path)

            output_filename = f"Reviewed_{site_name.replace(' ', '_')}_{unique_id}.xlsx"
            output_path = os.path.join(OUTPUT_FOLDER, output_filename)
            
            try:
                # ---------------------------------------------------------
                # NEW DATA INGESTION ENGINE
                # ---------------------------------------------------------
                nist_metrics = None
                cis_metrics = None # NEW initialization
                
                if upload_path.endswith('.txt'):
                    from cli_parser import PaloAltoCLIParser
                    with open(upload_path, 'r', encoding='utf-8', errors='ignore') as f:
                        raw_text = f.read()
                    
                    cli_engine = PaloAltoCLIParser(raw_text)
                    # Unpack the 3 variables now!
                    df, nist_metrics, cis_metrics = cli_engine.parse() 
                else:
                    df = pd.read_csv(upload_path) if upload_path.endswith('.csv') else pd.read_excel(upload_path)

                df = df.fillna('')
                
                # ---------------------------------------------------------
                # VENDOR VALIDATION & PROCESSING (Unchanged)
                # ---------------------------------------------------------
                is_palo = any(col in df.columns for col in ['Rule Usage Hit Count', 'Rule Usage Rule Usage', 'URL Category', 'Source Zone', 'Last Hit Date'])
                is_forti = any(col in df.columns for col in ['Policy', 'Security Profiles', 'Hit Count', 'From'])
                
                if is_palo:
                    required_pa = ['Name', 'Source Zone', 'Source Address', 'Destination Zone', 'Destination Address', 'Application', 'Service', 'Action', 'Profile', 'Options', 'Tags']
                    missing = [col for col in required_pa if col not in df.columns]
                    
                    if 'Rule Usage Hit Count' not in df.columns and 'Rule Usage Rule Usage' not in df.columns:
                        missing.append('Rule Usage Hit Count')
                    
                    if missing:
                        flash(f"Palo Alto Export Error - You are missing the following required columns: {', '.join(missing)}")
                        return redirect(request.url)
                        
                    # Hand both metrics off to the processor!
                    metrics, web_data = process_palo_alto_rules(df, output_path, site_name, nist_metrics=nist_metrics, cis_metrics=cis_metrics)
                    
                elif is_forti:
                    required_ft = ['Status', 'From', 'Source', 'To', 'Destination', 'Service', 'Action', 'Security Profiles', 'Log', 'Hit Count']
                    missing = [col for col in required_ft if col not in df.columns]
                    
                    if 'Name' not in df.columns and 'Policy' not in df.columns:
                        missing.append('Name (or Policy)')
                        
                    if missing:
                        flash(f"Fortinet Export Error - You are missing the following required columns: {', '.join(missing)}")
                        return redirect(request.url)
                        
                    metrics, web_data = process_fortinet_rules(df, output_path, site_name)
                    
                else:
                    flash("Unrecognized Format. Could not detect standard Palo Alto or Fortinet headers.")
                    return redirect(request.url)
                
                return render_template('summary.html', metrics=metrics, output_filename=output_filename, site_name=site_name, web_data=web_data)
                
            except Exception as e:
                flash(f"Processing Error: {str(e)}")
                return redirect(request.url)

    return render_template('index.html')

@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    return send_file(file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, port=5000)